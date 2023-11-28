# The input locations can be any directory or file path. However, the data directory must have a certain format.
# DEM file inside ../DEM, climate files inside ../Meteo
# Satellite files inside ../Satellite-data and soil files inside ../Soil
# The script can be altered to have different names of course.
# The climate files must include the date in YYYYMMDD format, much like the satellite files downloaded.

import os
import pandas as pd
from openpyxl import load_workbook

class pysebal_xl:
    def __init__(self, data_dir, xl_path, img_type):
        self.data_dir = data_dir
        self.xl_path = xl_path
        self.workbook = load_workbook(xl_path)
        self.img_type = img_type
    def find_location(self, folder):
        templist1 = []
        folder_path = os.path.join(self.data_dir, folder)
        for folder_obj in os.listdir(folder_path):
            data_folder_path = os.path.join(folder_path, folder_obj)
            templist1.append(data_folder_path)
        return templist1
    def sort_meteo(self, meteo_var, date_list, meteo_list):
        templist2 = []
        for date in date_list:
            for meteo_file in meteo_list:
                strings_to_search = [date, meteo_var]
                meteo_file_path = os.path.join(self.data_dir, "Meteo", meteo_file)
                if all(search_string in meteo_file for search_string in strings_to_search):
                    templist2.append(meteo_file_path)
        return templist2
    def write_xl(self, sheet_name, column_number, listname):
        worksheet = self.workbook[sheet_name]
        column_index = int(column_number)
        for index, value in enumerate(listname, start=2):
            worksheet.cell(row=index, column=column_index).value = value
        self.workbook.save(self.xl_path)
    def fill_xl(self):
        # Input locations
        satellite_data_list = self.find_location("Satellite_data")
        list_length = len(satellite_data_list)
        meteo_list = self.find_location("Meteo")
        soil_base_list = self.find_location("Soil")
        soil_list = [[obj] * list_length for obj in soil_base_list]
        dem_list = self.find_location("DEM") * list_length
        
        # List of output folder names to create separate output folders for each image and the output locations
        output_folder_names = [foldername.split('\\')[-1] for foldername in satellite_data_list]
        
        # Making output folders
        os.makedirs(os.path.join(self.data_dir, "SEBAL_Out"), exist_ok=True)
        for folder_name in output_folder_names:
            output_folder_path = os.path.join(self.data_dir, "SEBAL_Out", folder_name)
            os.makedirs(output_folder_path, exist_ok=True)
        output_list = self.find_location("SEBAL_Out")
        # Necessary input lists of meteo data and satellite names
        date_list = [obj[17:25] for obj in output_folder_names]
        sat_list  = [obj[0:4] for obj in output_folder_names]
        
        # List of meteo data locations
        Tair_inst_list = self.sort_meteo("Tair_inst", date_list, meteo_list)
        Tair_24_list = self.sort_meteo("Tair_24", date_list, meteo_list)
        Rh_inst_list = self.sort_meteo("Rh_inst", date_list, meteo_list)
        Rh_24_list = self.sort_meteo("Rh_24", date_list, meteo_list)
        Wind_inst_list = self.sort_meteo("Wind_inst", date_list, meteo_list)
        Wind_24_list = self.sort_meteo("Wind_24", date_list, meteo_list)
        SWdown_inst_list = self.sort_meteo("SWdown_inst", date_list, meteo_list)
        SWdown_24_list = self.sort_meteo("SWdown_24", date_list, meteo_list)
        
        # List of other parameters
        image_type_list = [self.img_type] * list_length
        zx_list = [2] * list_length
        method_radiation_list = [1] * list_length
        method_radiation_inst_list = [1] * list_length
        transm_24_list = [0.7] * list_length
        transm_inst_cloud_free_list = [0.75] * list_length
        obstacle_height_list = [0.4] * list_length
        depletion_factor_list = [0.5] * list_length     # 0.5 average value (0.2-0.3 for shallow and 0.4-0.6 for deep roots)
        luemax_list = [2.5] * list_length
        
        # Landsat 5, 7 and 8 parameters
        LT05 = [5, 1, 2, 5, 1, 3, 0.0065]
        LE07 = [7, 1, 2, 5, 1, 3, 0.0065]
        LC08 = [8, 2, 2, 5, 1, 3, 0.0065]
        LC09 = [9, 2, 2, 5, 1, 3, 0.0065]
        sat_dict = {'LT05': LT05, 'LE07': LE07, 'LC08': LC08}
        
        # Writing the locations and parameters
        # General input
        self.write_xl("General_Input", 2, satellite_data_list)
        self.write_xl("General_Input", 3, output_list)
        self.write_xl("General_Input", 4, image_type_list)
        self.write_xl("General_Input", 5, dem_list)
        
        # Meteo input
        self.write_xl("Meteo_Input", 2, Tair_inst_list)
        self.write_xl("Meteo_Input", 3, Tair_24_list)
        self.write_xl("Meteo_Input", 4, Rh_inst_list)
        self.write_xl("Meteo_Input", 5, Rh_24_list)
        self.write_xl("Meteo_Input", 6, zx_list)
        self.write_xl("Meteo_Input", 7, Wind_inst_list)
        self.write_xl("Meteo_Input", 8, Wind_24_list)
        self.write_xl("Meteo_Input", 9, method_radiation_list)
        self.write_xl("Meteo_Input", 10, SWdown_24_list)
        self.write_xl("Meteo_Input", 11, transm_24_list)
        self.write_xl("Meteo_Input", 12, method_radiation_inst_list)
        self.write_xl("Meteo_Input", 13, SWdown_inst_list)
        self.write_xl("Meteo_Input", 14, transm_inst_cloud_free_list)
        self.write_xl("Meteo_Input", 15, obstacle_height_list)
        
        # Soil input
        self.write_xl("Soil_Input", 2, soil_list[5])
        self.write_xl("Soil_Input", 3, soil_list[4])
        self.write_xl("Soil_Input", 4, soil_list[3])
        self.write_xl("Soil_Input", 5, soil_list[2])
        self.write_xl("Soil_Input", 6, soil_list[1])
        self.write_xl("Soil_Input", 7, soil_list[0])
        self.write_xl("Soil_Input", 8, depletion_factor_list)
        self.write_xl("Soil_Input", 9, luemax_list)     # C3/C4 crops
        
        if self.img_type == 1:
            # Landsat input
            self.write_xl("Landsat_Input", 2, output_folder_names)
            sheet = self.workbook["Landsat_Input"]
            for row in sheet.iter_rows(min_row=2, max_row=list_length+1, min_col=2, max_col=2):
                for cell in row:
                    sat_name = str(cell.value)[0:4]
                    if sat_name in sat_dict:
                        sat_data_list = sat_dict[sat_name]
                        
                        # Start from column C
                        column_index = 3
                        
                        # Write the data from the list to the corresponding columns
                        for sat_data in sat_data_list:
                            sheet.cell(row = cell.row, column = column_index, value = sat_data)
                            column_index += 1
                self.workbook.save(self.xl_path)
        
        else:
            # PROBAV and VIIRS input
            print("Support unavailable")
        
        print("Processing PySEBAL spreadsheet completed.")
